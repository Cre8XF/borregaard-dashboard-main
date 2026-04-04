"""
oppdater_dashboard.py — Borregaard Dashboard daglig oppdatering

Rogers morgenrutine:
  1. Eksporter 4 Excel-filer fra Butler/Jeeves
  2. Dobbeltklikk oppdater_dashboard.bat
  3. Ferdig
"""

import os
import re
import sys
import json
import shutil
import subprocess
from datetime import datetime

try:
    import pandas as pd
except ImportError:
    print("❌ pandas er ikke installert. Kjør: pip install pandas openpyxl")
    sys.exit(1)

# ── Filstier ────────────────────────────────────────────────────────────────
BASE     = r"C:\Users\ROGSOR0319\_Datahub\Excel-eksporter"
DAGLIG   = os.path.join(BASE, "01-Daglig")
UKENTLIG = os.path.join(BASE, "02_Ukentlig")
SJELDEN  = os.path.join(BASE, "03-Sjelden")

required = {
    "Master_Artikkelstatus.xlsx":   DAGLIG,
    "Master.xlsx":                  DAGLIG,
    "Ordrer_Jeeves.xlsx":           DAGLIG,
    "bestillinger.xlsx":            DAGLIG,
    "Inventeringshistorikk.xlsx":   DAGLIG,      # NY
    "dagsomsetning.xlsx":           DAGLIG,      # NY
    "Orderingang.xlsx":             DAGLIG,      # FASE 10.x
    "SA-Nummer.xlsx":               SJELDEN,
    "leverandører.xlsx":            SJELDEN,
    "Analyse_Lagerplan.xlsx":       UKENTLIG,
}

# Prisliste er valgfri — dashbordet fungerer uten den
PRISLISTE_PATH = r"C:\Users\ROGSOR0319\_Datahub\Excel-eksporter\03-Sjelden\20260319_Borregaard_prisliste.xlsx"

# Ordrestockanalys er valgfri — periodisk fil, legg i 03-Sjelden ved rapportbehov
ORDRESTOCKANALYS_PATH = os.path.join(SJELDEN, "Ordrestockanalys.xlsx")

# Lavverdi-telleliste er valgfri
LAVVERDI_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Lavverdi_Telleliste_2026.xlsx")

# data (7).xlsx sjekkes separat pga. rename
data7_src = os.path.join(SJELDEN, "data (7).xlsx")

# ── Steg 0: Sjekk at alle filer er på plass ─────────────────────────────────
print("Sjekker påkrevde filer...")
mangler = False
for fil, mappe in required.items():
    full_sti = os.path.join(mappe, fil)
    if not os.path.exists(full_sti):
        print(f"❌ Mangler fil: {fil}")
        print(f"   Forventet sti: {full_sti}")
        mangler = True
    else:
        print(f"✅ {fil}")

if not os.path.exists(data7_src):
    print(f"❌ Mangler fil: data (7).xlsx")
    print(f"   Forventet sti: {data7_src}")
    mangler = True
else:
    print(f"✅ data (7).xlsx")

if mangler:
    print("\nEksporter manglende filer fra Butler/Jeeves og prøv igjen.")
    sys.exit(1)

try:
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # ── Steg 1: Kopier filer til prosjektmappen ──────────────────────────────
    print("\n[1/5] Kopierer Excel-filer til prosjektmappen...")
    for fil, mappe in required.items():
        src = os.path.join(mappe, fil)
        dst = os.path.join(script_dir, fil)
        shutil.copy2(src, dst)
        print(f"  → {fil}")

    # data (7).xlsx kopieres med nytt navn
    shutil.copy2(data7_src, os.path.join(script_dir, "data_7.xlsx"))
    print(f"  → data (7).xlsx  (lagret som data_7.xlsx)")
    print("✅ Filer kopiert")

    # ── Steg 2: Kjør generate_masterV2.py ────────────────────────────────────
    print("\n[2/5] Genererer Borregaard_SA_Master_v2.xlsx...")
    subprocess.run([sys.executable, "generate_masterV2.py"], check=True,
                   cwd=script_dir)
    print("✅ MV2 generert")

    # ── Steg 3: Les Excel-filer og bygg JSON ─────────────────────────────────
    print("\n[3/5] Konverterer Excel til JSON...")

    master = pd.read_excel(
        os.path.join(script_dir, "Borregaard_SA_Master_v2.xlsx"),
        sheet_name="SA-Oversikt",
        dtype=str
    ).fillna("")

    orders = pd.read_excel(
        os.path.join(script_dir, "Ordrer_Jeeves.xlsx"), dtype=str
    ).fillna("")

    best = pd.read_excel(
        os.path.join(script_dir, "bestillinger.xlsx"), dtype=str
    ).fillna("")

    # ── Orderingang — DG% per ordrelinje (FASE 10.x) ─────────────────────────
    oi_records = []
    try:
        oi_df = pd.read_excel(
            os.path.join(script_dir, "Orderingang.xlsx"),
            dtype=str
        ).fillna("")

        # Filtrer kun Borregaard - TOOLS
        oi_df = oi_df[oi_df["LstK"].str.strip() == "Borregaard - TOOLS"].copy()

        def parse_no(val):
            """Parse norsk tallformat: '47 264,50' → 47264.5"""
            try:
                return float(str(val).replace(" ", "").replace(",", "."))
            except (ValueError, TypeError):
                return None

        oi_map = {}  # key = "ordrenr|artnr"
        for _, row in oi_df.iterrows():
            ordrenr   = str(row.get("OrderNr", "")).strip()
            artnr_raw = str(row.get("Artikelnr", "")).strip()
            artnr     = re.sub(r'-\w{1,3}$', '', artnr_raw)
            if not ordrenr or not artnr:
                continue
            radbidrag_pct = parse_no(row.get("Radbidrag i %", ""))
            radbidr       = parse_no(row.get("Radbidr i basvaluta", "")) or 0
            radverdi      = parse_no(row.get("Radvärde i basvaluta", "")) or 0
            prisval       = parse_no(row.get("PrisVal", "")) or 0
            if radbidrag_pct is None:
                continue
            key = f"{ordrenr}|{artnr}"
            if key not in oi_map:
                oi_map[key] = {"ordrenr": ordrenr, "artnr": artnr,
                               "radbidr": 0, "radverdi": 0,
                               "prisval_sum": 0, "prisval_count": 0}
            oi_map[key]["radbidr"]       += radbidr
            oi_map[key]["radverdi"]      += radverdi
            oi_map[key]["prisval_sum"]   += prisval
            oi_map[key]["prisval_count"] += 1

        oi_records = []
        for key, v in oi_map.items():
            dg = (v["radbidr"] / v["radverdi"] * 100) if v["radverdi"] > 0 else 0
            prisval_snitt = (v["prisval_sum"] / v["prisval_count"]) if v["prisval_count"] > 0 else 0
            oi_records.append({
                "ordrenr":   v["ordrenr"],
                "artnr":     v["artnr"],
                "radbidrag": round(dg, 3),
                "radbidr":   round(v["radbidr"], 2),
                "radverdi":  round(v["radverdi"], 2),
                "prisval":   round(prisval_snitt, 2),
            })

        print(f"✅ Orderingang lastet ({len(oi_records)} linjer med DG%)")
    except Exception as oi_err:
        print(f"⚠️  Orderingang ikke tilgjengelig (fortsetter uten): {oi_err}")

    # ── Dagsomsetning (valgfri) ─────────────────────────────────────────────────
    salg_records = []
    try:
        salg_df = pd.read_excel(
            os.path.join(script_dir, "dagsomsetning.xlsx"), dtype=str
        ).fillna("")
        salg_records = salg_df.to_dict(orient="records")
        print(f"✅ Dagsomsetning lastet ({len(salg_records)} rader)")
    except Exception as salg_err:
        print(f"⚠️  dagsomsetning.xlsx ikke tilgjengelig (fortsetter uten): {salg_err}")

    # ── Prisliste (valgfri) ───────────────────────────────────────────────────
    pris_records = []
    try:
        pris_df = pd.read_excel(PRISLISTE_PATH, header=4, dtype=str)
        pris_df.columns = [str(c).strip() for c in pris_df.columns]

        # Konverter priskolonner (komma → punktum)
        for col in ['q_replacement_value', 'Ny pris', 'Ny DG']:
            if col in pris_df.columns:
                pris_df[col] = pd.to_numeric(
                    pris_df[col].astype(str).str.replace(',', '.').str.replace(' ', ''),
                    errors='coerce'
                ).fillna(0)

        if 'artlistpris' in pris_df.columns:
            pris_df['artlistpris'] = pd.to_numeric(
                pris_df['artlistpris'].astype(str).str.replace(',', '.').str.replace(' ', ''),
                errors='coerce'
            ).fillna(0)

        pris_cols = ['artnr', 'SA-Nummer', 'q_replacement_value', 'artlistpris',
                     'Ny pris', 'Ny DG', 'Status', 'Anbefaling']
        pris_cols = [c for c in pris_cols if c in pris_df.columns]
        pris_records = pris_df[pris_cols].fillna('').to_dict(orient="records")
        print(f"✅ Prisliste lastet ({len(pris_records)} rader)")
    except Exception as pris_err:
        print(f"⚠️  Prisliste ikke tilgjengelig (fortsetter uten): {pris_err}")

    # Sjekk prisliste for kalkylpris-fallback
    if os.path.exists(PRISLISTE_PATH):
        print(f'✅ Prisliste funnet — kalkylpris fallback aktivert')
    else:
        print(f'⚠️  Prisliste ikke funnet — kalkylpris kun fra Master_Artikkelstatus')

    # ── Ordrestockanalys (valgfri, periodisk) ──────────────────────────────────
    ordrestock_records = []
    try:
        if os.path.exists(ORDRESTOCKANALYS_PATH):
            print("Leser Ordrestockanalys.xlsx...")
            os_df = pd.read_excel(ORDRESTOCKANALYS_PATH, dtype=str)
            os_df.columns = [str(c).strip() for c in os_df.columns]

            # Filtrer kun Borregaard (424186 + 449930) og lager 3018
            if 'Företagsnr' in os_df.columns:
                os_df = os_df[os_df['Företagsnr'].isin(['424186', '449930'])]
            if 'LstK' in os_df.columns:
                os_df = os_df[os_df['LstK'].astype(str).str.strip().isin(['3018', 'Borregaard - TOOLS'])]

            def clean_num(s):
                """Rens komma-desimal og mellomrom-tusenskille til float."""
                if pd.isna(s) or str(s).strip() == '':
                    return ''
                return str(s).replace('\xa0', '').replace(' ', '').replace(',', '.')

            # Rens tallkolonner
            for col in ['OrdRadAnt', 'Radvärde i basvaluta', 'Radbidrag i %',
                        'Radbidr i basvaluta', 'Inköpspris']:
                if col in os_df.columns:
                    os_df[col] = os_df[col].apply(clean_num)

            # Behold kun relevante kolonner
            keep_cols = [
                'Artikelnr', 'OrdDtm', 'OrdRadAnt', 'Radvärde i basvaluta',
                'Radbidrag i %', 'Radbidr i basvaluta', 'Inköpspris',
                'Företagsnr', 'LstK', 'FaktDat', 'Kunds artikelnummer',
                'LevPlFtgKod', 'Artikelbeskrivning'
            ]
            keep_cols = [c for c in keep_cols if c in os_df.columns]
            os_df = os_df[keep_cols].fillna('')

            ordrestock_records = os_df.to_dict(orient='records')
            aapne = os_df[os_df['FaktDat'] == ''].shape[0] if 'FaktDat' in os_df.columns else '?'
            print(f"✅ Ordrestockanalys lastet ({len(ordrestock_records)} rader, {aapne} åpne ordrer)")
        else:
            print("⚠️  Ordrestockanalys.xlsx ikke funnet i 03-Sjelden (fortsetter uten)")
    except Exception as os_err:
        print(f"⚠️  Ordrestockanalys ikke tilgjengelig (fortsetter uten): {os_err}")

    # ── DG-kontroll: Orderingang.xlsx (FASE 9.x) ─────────────────────────────
    print("\nLeser Orderingang.xlsx for DG-kontroll...")
    ORDERINGANG_PATH = os.path.join(DAGLIG, "Orderingang.xlsx")
    dg_kontroll = {}
    try:
        from datetime import timedelta
        og_raw = pd.read_excel(ORDERINGANG_PATH, header=0, dtype=str)
        og_raw.columns = [str(c).strip() for c in og_raw.columns]

        col_art_nr = "Artikelnr"
        col_beskr  = "Artikelbeskrivning"
        col_pris   = "PrisVal"
        col_ksv    = "Radbidr i basvaluta"
        col_dg     = "Radbidrag i %"
        col_dato   = "OrdDtm"

        # Filtrer kun Invoiced-rader
        og = og_raw[og_raw["OrdRdSt"].astype(str).str.strip() == "Invoiced"].copy()

        # Parse dato YYMMDD → datetime
        def parse_yymmdd(val):
            try:
                s = str(int(float(str(val)))).zfill(6)
                return datetime(2000 + int(s[0:2]), int(s[2:4]), int(s[4:6]))
            except Exception:
                return None

        og['_dato'] = og[col_dato].apply(parse_yymmdd)
        og = og.dropna(subset=['_dato'])

        # Konverter numeriske kolonner (komma → punktum)
        for c in [col_pris, col_ksv, col_dg]:
            og[c] = pd.to_numeric(
                og[c].astype(str).str.replace(',', '.').str.replace(' ', ''),
                errors='coerce'
            ).fillna(0.0)

        # Filtrer siste 12 måneder (365 dager)
        cutoff_12m = datetime.now() - timedelta(days=365)
        og_12m = og[og['_dato'] >= cutoff_12m].copy()

        # Grupper per Artikelnr — hent siste ordre og 12-mnd snitt
        grp12_map = {
            str(art_nr).strip(): grp
            for art_nr, grp in og_12m.groupby(col_art_nr)
            if str(art_nr).strip()
        }

        for art_nr, grp_all in og.groupby(col_art_nr):
            art_nr = str(art_nr).strip()
            if not art_nr:
                continue

            # Seneste ordre (høyest dato)
            latest = grp_all.loc[grp_all['_dato'].idxmax()]

            grp12 = grp12_map.get(art_nr, pd.DataFrame())
            antall_12m = len(grp12)

            dg_snitt_12mnd = round(float(grp12[col_dg].mean()), 1) if antall_12m > 0 \
                             else round(float(latest[col_dg]), 1)
            dg_siste = round(float(latest[col_dg]), 1)
            dg_avvik = round(dg_siste - dg_snitt_12mnd, 1)

            # dg_trend_3: snitt DG% på de 3 siste ordrene (vedvarende trend-indikator)
            if antall_12m >= 3:
                siste3 = grp12.nlargest(3, '_dato')
                dg_trend_3 = round(float(siste3[col_dg].mean()), 1)
            elif antall_12m > 0:
                dg_trend_3 = round(float(grp12[col_dg].mean()), 1)
            else:
                dg_trend_3 = dg_siste

            dg_kontroll[art_nr] = {
                "beskrivelse":      str(latest[col_beskr]).strip(),
                "dg_snitt_12mnd":   dg_snitt_12mnd,
                "dg_siste":         dg_siste,
                "dg_trend_3":       dg_trend_3,
                "dg_avvik":         dg_avvik,
                "siste_pris":       round(float(latest[col_pris]), 2),
                "siste_ksv":        round(float(latest[col_ksv]), 2),
                "siste_ordredato":  latest['_dato'].strftime('%Y-%m-%d'),
                "antall_ordrer_12mnd": antall_12m,
            }

        print(f"✅ DG-kontroll: {len(dg_kontroll)} artikler analysert")

    except FileNotFoundError:
        print(f"⚠️  Orderingang.xlsx ikke funnet — dgKontroll satt til tom dict")
    except Exception as dg_err:
        print(f"⚠️  DG-kontroll feil (fortsetter uten): {dg_err}")

    # ── Vedlikeholdsstopp: faktisk forbrukshistorikk (FASE 10.x) ─────────────
    print("\nBeregner vedlikeholdsstopp-data fra Orderingang...")
    FOCUS_UKER = [16, 42]
    VINDU_UKER_MAP = {
        16: [14, 15, 16, 17, 18],
        42: [40, 41, 42, 43, 44],
    }
    vedlikeholdsstopp = {"uke16": {}, "uke42": {}}
    try:
        # Bygg oppslag: OrderNr -> Delivery location ID fra Ordrer_Jeeves
        jeeves_map_vs = {}
        for _, jrow in orders.iterrows():
            order_nr = str(jrow.get('Order number', '')).strip()
            delivery = str(jrow.get('Delivery location ID', '')).strip()
            if order_nr and delivery:
                jeeves_map_vs[order_nr] = delivery

        # Les Orderingang (gjenbruk ORDERINGANG_PATH fra DG-kontroll)
        og_vs = pd.read_excel(ORDERINGANG_PATH, header=0, dtype=str)
        og_vs.columns = [str(c).strip() for c in og_vs.columns]

        col_ordernr_vs = "OrderNr"
        col_artnr_vs   = "Artikelnr"
        col_qty_vs     = "OrdRadAnt"
        col_dato_vs    = "OrdDtm"

        # Filtrer kun Invoiced-rader
        og_vs = og_vs[og_vs["OrdRdSt"].astype(str).str.strip() == "Invoiced"].copy()

        # Parse dato YYMMDD → datetime
        def parse_yymmdd_vs(val):
            try:
                s = str(int(float(str(val)))).zfill(6)
                return datetime(2000 + int(s[0:2]), int(s[2:4]), int(s[4:6]))
            except Exception:
                return None

        og_vs['_dato'] = og_vs[col_dato_vs].apply(parse_yymmdd_vs)
        og_vs = og_vs.dropna(subset=['_dato'])

        # Konverter qty
        og_vs[col_qty_vs] = pd.to_numeric(
            og_vs[col_qty_vs].astype(str).str.replace(',', '.').str.replace(' ', ''),
            errors='coerce'
        ).fillna(0.0)

        # Berik med leveringssted og ISO-uke
        og_vs['_delivery_id'] = og_vs[col_ordernr_vs].astype(str).str.strip().apply(
            lambda x: jeeves_map_vs.get(x, '')
        )
        og_vs['_uke']   = og_vs['_dato'].apply(lambda d: d.isocalendar()[1])
        og_vs['_ar']    = og_vs['_dato'].apply(lambda d: d.year)
        og_vs['_artnr'] = og_vs[col_artnr_vs].astype(str).str.strip()

        for focus_uke in FOCUS_UKER:
            uke_key    = f"uke{focus_uke}"
            vindu_uker = VINDU_UKER_MAP[focus_uke]

            og_focus = og_vs[og_vs['_uke'] == focus_uke]
            og_vindu = og_vs[og_vs['_uke'].isin(vindu_uker)]

            artnrs_in_focus = [a for a in og_focus['_artnr'].unique() if a]

            for artnr in artnrs_in_focus:
                focus_rader = og_focus[og_focus['_artnr'] == artnr]
                vindu_rader = og_vindu[og_vindu['_artnr'] == artnr]

                # Historisk snitt i fokusuke (snitt over år)
                focus_by_year = focus_rader.groupby('_ar')[col_qty_vs].sum()
                antall_ar = len(focus_by_year)
                historisk_snitt_focus = round(float(focus_by_year.mean()), 1) if antall_ar > 0 else 0.0

                # Historisk snitt i vindu (snitt over år)
                vindu_by_year = vindu_rader.groupby('_ar')[col_qty_vs].sum()
                historisk_snitt_vindu = round(float(vindu_by_year.mean()), 1) if len(vindu_by_year) > 0 else 0.0

                # Per leveringssted i fokusuke
                per_leveringssted = {}
                for _, row in focus_rader.iterrows():
                    dlid = str(row['_delivery_id']).strip()
                    if not dlid:
                        continue
                    per_leveringssted[dlid] = round(
                        per_leveringssted.get(dlid, 0.0) + float(row[col_qty_vs]), 1
                    )

                # Sesongtype
                if antall_ar == 1:
                    sesong_type = "engang"
                elif historisk_snitt_vindu > 0:
                    avg_vindu_uke = historisk_snitt_vindu / len(vindu_uker)
                    ratio = historisk_snitt_focus / avg_vindu_uke if avg_vindu_uke > 0 else 1.0
                    sesong_type = "spike" if ratio > 1.5 else "jevn"
                else:
                    sesong_type = "jevn"

                anbefalt_innkjop = round(historisk_snitt_focus * 1.2)
                beskrivelse = dg_kontroll.get(artnr, {}).get('beskrivelse', '')

                vedlikeholdsstopp[uke_key][artnr] = {
                    "beskrivelse":           beskrivelse,
                    "historisk_snitt_focus": historisk_snitt_focus,
                    "historisk_snitt_vindu": historisk_snitt_vindu,
                    "antall_ar_med_data":    antall_ar,
                    "per_leveringssted":     per_leveringssted,
                    "anbefalt_innkjop":      anbefalt_innkjop,
                    "sesong_type":           sesong_type,
                }

        tot16 = len(vedlikeholdsstopp['uke16'])
        tot42 = len(vedlikeholdsstopp['uke42'])
        print(f"✅ Vedlikeholdsstopp: uke16={tot16} artikler, uke42={tot42} artikler")

    except FileNotFoundError:
        print(f"⚠️  Orderingang.xlsx ikke funnet — vedlikeholdsstopp satt til tom dict")
    except Exception as vs_err:
        print(f"⚠️  Vedlikeholdsstopp-feil (fortsetter uten): {vs_err}")

    # ── Lavverdi-telleliste (FASE 11.0) ──────────────────────────────────────
    lavverdi_rows = []
    try:
        from openpyxl import load_workbook
        lv_wb = load_workbook(LAVVERDI_PATH, read_only=True, data_only=True)
        lv_ws = lv_wb.active
        for row in lv_ws.iter_rows(min_row=5, values_only=True):
            if not row[0]:
                continue  # hopp over tomme rader
            lavverdi_rows.append({
                'lokasjon':    str(row[0] or '').strip(),
                'tools_artnr': str(row[1] or '').strip(),
                'sa_nummer':   str(row[2] or '').strip(),
                'beskrivelse': str(row[3] or '').strip(),
                'saldo':       row[4] or 0,
                'kalkylpris':  row[5] or 0,
                'est_verdi':   row[6] or 0,
                'sist_telt':   str(row[7] or '').strip(),
            })
        lv_wb.close()
        print(f"✅ Lavverdi-telleliste: {len(lavverdi_rows)} artikler")
    except FileNotFoundError:
        print(f"⚠️  Lavverdi_Telleliste_2026.xlsx ikke funnet — lavverdiListe satt til tom liste")
    except Exception as lv_err:
        print(f"⚠️  Lavverdi-telleliste feil (fortsetter uten): {lv_err}")

    # ── Bevegelse: siste salg + siste innlevering (FASE 11.x) ────────────────
    print("\nBeregner bevegelsesdata...")
    bevegelse = {}
    try:
        siste_salg_map = {}
        try:
            # 'og' og 'col_art_nr' er definert i DG-kontroll-blokken ovenfor
            for art_nr, grp in og.groupby(col_art_nr):
                art_nr_str = str(art_nr).strip()
                if not art_nr_str:
                    continue
                siste_dato = grp['_dato'].max()
                siste_salg_map[art_nr_str] = siste_dato.strftime('%Y-%m-%d')
        except NameError:
            print("⚠️  Orderingang ikke tilgjengelig for bevegelse (DG-kontroll feilet) — siste_salg settes til null")

        # Siste innlevering fra bestillinger.xlsx — kol [7]=Artikelnr, [115]=InlevDat (YYMMDD)
        inlev_map = {}
        if len(best.columns) > 115:
            for _, row in best.iterrows():
                artnr     = str(row.iloc[7]   or '').strip()
                inlev_raw = str(row.iloc[115] or '').strip()
                if artnr and len(inlev_raw) == 6:
                    try:
                        dato = datetime.strptime('20' + inlev_raw, '%Y%m%d')
                        iso  = dato.strftime('%Y-%m-%d')
                        if artnr not in inlev_map or iso > inlev_map[artnr]:
                            inlev_map[artnr] = iso
                    except Exception:
                        pass
        else:
            print("⚠️  bestillinger.xlsx har færre enn 116 kolonner — InlevDat ikke tilgjengelig")

        # Slå sammen til bevegelse-objekt
        alle_artnr = set(siste_salg_map.keys()) | set(inlev_map.keys())
        for artnr in alle_artnr:
            salg  = siste_salg_map.get(artnr)
            inlev = inlev_map.get(artnr)
            kandidater = [d for d in [salg, inlev] if d]
            siste = max(kandidater) if kandidater else None
            bevegelse[artnr] = {
                'siste_salg':      salg,
                'siste_inlev':     inlev,
                'siste_bevegelse': siste,
            }
        print(f"✅ Bevegelse: {len(bevegelse)} artikler ({len(siste_salg_map)} med salg, {len(inlev_map)} med innlev.)")
    except Exception as bev_err:
        print(f"⚠️  Bevegelse-beregning feil (fortsetter uten): {bev_err}")

    # ── Varetelling 2026: beregn tellingsomfang og telt (FASE 8.1b — Butler-logikk) ──
    print("\n[Varetelling] Beregner tellingsomfang og telt i 2026 (Butler-logikk)...")

    try:
        # ── Krit 1: artikler med saldo nå ────────────────────────────────────────
        master_full = pd.read_excel(os.path.join(script_dir, "Master.xlsx"), dtype=str)
        master_full.columns = master_full.columns.str.strip()
        _lstk_cols = [i for i, c in enumerate(master_full.columns) if c == 'LstK']
        _lstk_idx = _lstk_cols[0] if _lstk_cols else None
        if _lstk_idx is None:
            raise ValueError("Finner ikke LstK-kolonne i Master.xlsx")
        master_3018 = master_full[master_full.iloc[:, _lstk_idx].astype(str).str.strip() == '3018'].copy()
        master_3018['_saldo'] = pd.to_numeric(master_3018['Lagersaldo (hylla)'], errors='coerce').fillna(0)
        master_3018['_artnr'] = master_3018['Artikelnr'].astype(str).str.strip()

        arts_krit1 = set(master_3018[master_3018['_saldo'] > 0]['_artnr'])

        # ── Krit 2: solgt i 2026 (Ordrer_Jeeves) ─────────────────────────────────
        ordrer_df = pd.read_excel(os.path.join(script_dir, "Ordrer_Jeeves.xlsx"), dtype=str)
        ordrer_df.columns = ordrer_df.columns.str.strip()
        ordrer_2026 = ordrer_df[ordrer_df['Date'].astype(str).str.startswith('2026')]
        arts_krit2 = set(ordrer_2026['Item ID'].astype(str).str.strip())

        # ── Krit 3: mottatt i 2026 (bestillinger InlevDat starter med '260') ─────
        best_df = pd.read_excel(os.path.join(script_dir, "bestillinger.xlsx"), dtype=str)
        best_df.columns = best_df.columns.str.strip()
        best_3018 = best_df[best_df['LstK'].astype(str).str.strip() == '3018'].copy()
        best_3018['_inlev'] = best_3018['InlevDat'].astype(str).str.strip()
        arts_krit3 = set(best_3018[best_3018['_inlev'].str.startswith('260')]['Artikelnr'].astype(str).str.strip())

        # ── Omfang = union av alle 3 kriterier ────────────────────────────────────
        omfang_arts = arts_krit1 | arts_krit2 | arts_krit3

        # ── Telt = artikler i omfang som finnes i Inventeringshistorikk 2026 ──────
        inv_path = os.path.join(DAGLIG, "Inventeringshistorikk.xlsx")
        inv_df = pd.read_excel(inv_path, dtype=str)
        inv_df.columns = inv_df.columns.str.strip()
        credt_col = next((c for c in inv_df.columns if c.strip() == 'CreDt'), None)
        artnr_col = next((c for c in inv_df.columns if c.strip() == 'Artikelnr'), None)

        if credt_col and artnr_col:
            inv_2026 = inv_df[inv_df[credt_col].astype(str).str.startswith('26')]
            inv_arts_2026 = set(inv_2026[artnr_col].astype(str).str.strip())
        else:
            print("  ⚠️ Fant ikke CreDt/Artikelnr i Inventeringshistorikk")
            inv_arts_2026 = set()

        telt_arts = inv_arts_2026 & omfang_arts

        varetelling_meta = {
            "omfang":         len(omfang_arts),
            "antall_telt":    len(telt_arts),
            "prosent_telt":   round(len(telt_arts) / len(omfang_arts) * 100, 1) if omfang_arts else 0,
            "sist_oppdatert": datetime.now().strftime("%Y-%m-%d"),
            "logikk":         "Butler: saldo>0 OR solgt 2026 OR mottatt 2026"
        }

        print(f"  Krit 1 (saldo > 0):      {len(arts_krit1)}")
        print(f"  Krit 2 (solgt i 2026):   {len(arts_krit2)} ({len(arts_krit2 - arts_krit1)} nye)")
        print(f"  Krit 3 (mottatt i 2026): {len(arts_krit3)} ({len(arts_krit3 - arts_krit1 - arts_krit2)} nye)")
        print(f"  Omfang totalt:           {varetelling_meta['omfang']}")
        print(f"  Telt i 2026:             {varetelling_meta['antall_telt']}")
        print(f"  Prosent:                 {varetelling_meta['prosent_telt']}%")

    except Exception as vt_err:
        print(f"  ⚠️ Varetelling-beregning feilet (fortsetter uten): {vt_err}")
        varetelling_meta = {
            "omfang": 0,
            "antall_telt": 0,
            "prosent_telt": 0,
            "sist_oppdatert": "",
            "logikk": "feil"
        }

    # ── Utskutte lager (FASE 12) ─────────────────────────────────────────────────
    UTSKUTTE_MAP = {
        "424186-2": "ØST: Spriten",
        "424186-3": "VEST: Cellulose",
        "424186-4": "Kokeri",
        "424186-5": "ALVA",
        "424186-6": "SENTRALV.",
        "424186-7": "Grace",
    }
    UTSKUTTE_RAMPE = {
        "424186-2": "R-229",
        "424186-3": "R-275",
        "424186-4": "R-125",
        "424186-5": "R-265",
        "424186-6": "R-156",
        "424186-7": "",
    }

    utskutte_df = orders[orders["Delivery location ID"].isin(UTSKUTTE_MAP.keys())].copy()
    utskutte_df["lokasjon"]       = utskutte_df["Delivery location ID"].map(UTSKUTTE_MAP)
    utskutte_df["rampe"]          = utskutte_df["Delivery location ID"].map(UTSKUTTE_RAMPE)
    utskutte_df["deliv_id"]       = utskutte_df["Delivery location ID"]
    utskutte_df["Date"]           = pd.to_datetime(utskutte_df["Date"], errors="coerce")
    utskutte_df["Delivered value"]    = pd.to_numeric(utskutte_df["Delivered value"], errors="coerce").fillna(0)
    utskutte_df["Delivered quantity"] = pd.to_numeric(utskutte_df["Delivered quantity"], errors="coerce").fillna(0)

    # Records for JS (alle ordrelinjer, utskutte lager kun)
    utskutte_records = []
    for _, row in utskutte_df.iterrows():
        if pd.isna(row["Date"]):
            continue
        utskutte_records.append({
            "dato":      row["Date"].strftime("%Y-%m-%d"),
            "deliv_id":  row["Delivery location ID"],
            "lokasjon":  row["lokasjon"],
            "rampe":     row["rampe"],
            "item_id":   str(row.get("Item ID", "")),
            "item":      str(row.get("Item", "")),
            "verdi":     float(row["Delivered value"]),
            "antall":    float(row["Delivered quantity"]),
            "ordre_nr":  str(row.get("Order number", "")),
        })

    print(f"  Utskutte lager:          {len(utskutte_records)} ordrelinjer, {utskutte_df['Delivery location ID'].nunique()} lokasjoner")

    data = {
        "generert":           datetime.now().strftime("%Y-%m-%d %H:%M"),
        "master":             master.to_dict(orient="records"),
        "orders":             orders.to_dict(orient="records"),
        "bestillinger":       best.to_dict(orient="records"),
        "prisliste":          pris_records,         # FASE 9.0
        "dgKontroll":         dg_kontroll,           # FASE 9.x
        "vedlikeholdsstopp":  vedlikeholdsstopp,     # FASE 10.x
        "lavverdiListe":      lavverdi_rows,          # FASE 11.0
        "bevegelse":          bevegelse,              # FASE 11.x
        "varetelling_meta":   varetelling_meta,       # FASE 8.1
        "ordrestockanalys":   ordrestock_records,     # FASE 9.1 — valgfri, periodisk
        "dagsomsetning":      salg_records,           # NY
        "orderingang":        oi_records,             # FASE 10.x
        "utskutteLager":      utskutte_records,       # FASE 12
    }

    os.makedirs(os.path.join(script_dir, "data"), exist_ok=True)
    json_path = os.path.join(script_dir, "data", "dashboard-data.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)

    size_kb = os.path.getsize(json_path) / 1024
    print(f"✅ JSON generert ({size_kb:.0f} KB, {len(master)} artikler)")

    # ── Steg 4: Git push ──────────────────────────────────────────────────────
    print("\n[4/5] Pusher til GitHub...")
    subprocess.run(["git", "add", "data/dashboard-data.json"], check=True,
                   cwd=script_dir)
    subprocess.run(
        ["git", "commit", "-m",
         f"Data oppdatert {datetime.now().strftime('%Y-%m-%d %H:%M')}"],
        check=True, cwd=script_dir
    )
    subprocess.run(["git", "push"], check=True, cwd=script_dir)
    print("✅ Pushet til GitHub")

    # ── Steg 5: Ferdig ───────────────────────────────────────────────────────
    print("\n[5/5] Netlify redeployer automatisk (~10 sekunder)")
    print("\n🎉 Dashboard oppdatert! Kollega kan åpne linken nå.")

except subprocess.CalledProcessError as e:
    print(f"\n❌ Kommandofeil: {e}")
    print("   Sjekk at git er konfigurert og at du har tilgang til repoet.")
    sys.exit(1)
except Exception as e:
    print(f"\n❌ Uventet feil: {e}")
    sys.exit(1)