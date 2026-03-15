#!/usr/bin/env python3
"""
generate_masterV2.py — Genererer Borregaard_SA_Master_v2.xlsx  (FASE 7.2)

Kildefiler (alle i samme mappe som dette scriptet):
  1. data_7.xlsx                — PRIMÆR: SA-nr, Lokasjon, VareStatus, ErsattsAvArtNr
  2. Master_Artikkelstatus.xlsx — lagerstatus, kalkylpris, Artikelstatus
  3. Analyse_Lagerplan.xlsx     — BP, EOK, Maxlager, leverandørnummer
  4. Master.xlsx                — ErsattsAvArtNr (fallback), InvDat (sist telt)
  5. leverandører.xlsx          — LevLedTid, Transportdagar (nøkkel: Företagsnr)
  6. SA-Nummer.xlsx             — fallback SA-nr for artikler ikke i data_7 (IKKE lokasjon)

Output:
  Borregaard_SA_Master_v2.xlsx  — 33 kolonner, ark "SA-Oversikt"

Endringer:
  - Selvpekende erstatninger filtreres ut (ErsattsAvArtNr == Tools_ArtNr)
  - Alle lokasjoner fra data_7 aksepteres (LOK_PATTERN fjernet — BORSKUFF, LAGER1 osv. er gyldige)
  - Ordrer_Jeeves.xlsx kobles inn for salgshistorikk
"""

import os
import openpyxl
from openpyxl import Workbook

SCRIPT_DIR       = os.path.dirname(os.path.abspath(__file__))

DATA7_FILE       = os.path.join(SCRIPT_DIR, 'data_7.xlsx')
MASTER_FILE      = os.path.join(SCRIPT_DIR, 'Master_Artikkelstatus.xlsx')
MASTER_FULL_FILE = os.path.join(SCRIPT_DIR, 'Master.xlsx')
LAGERPLAN_FILE   = os.path.join(SCRIPT_DIR, 'Analyse_Lagerplan.xlsx')
LEVERANDOR_FILE  = os.path.join(SCRIPT_DIR, 'leverandører.xlsx')
SA_FILE          = os.path.join(SCRIPT_DIR, 'SA-Nummer.xlsx')
ORDRER_FILE      = os.path.join(SCRIPT_DIR, 'Ordrer_Jeeves.xlsx')
OUTPUT_FILE      = os.path.join(SCRIPT_DIR, 'Borregaard_SA_Master_v2.xlsx')


# ── Hjelpefunksjoner ──────────────────────────────────────────────────────────

def read_sheet(filepath, sheet_name=None):
    """Les ark til liste av dict {kolonnenavn: verdi}."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active if sheet_name is None else wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    headers = list(rows[0])
    data = []
    for raw_row in rows[1:]:
        if all(v is None for v in raw_row):
            continue
        d = {}
        seen = {}
        for i, h in enumerate(headers):
            if h is None:
                continue
            key = str(h).strip()
            if key in seen:
                seen[key] += 1
                key = f"{key}_{seen[key]}"
            else:
                seen[key] = 1
            d[key] = raw_row[i] if i < len(raw_row) else None
        data.append(d)
    return data


def val(d, *keys, default=''):
    """Hent første ikke-tomme verdi fra dict d for en av keys."""
    for k in keys:
        v = d.get(k)
        if v is not None and str(v).strip() != '':
            return v
    return default


def format_invdat(raw):
    """Konverter InvDat (int eller str YYYYMMDD) til streng 'YYYYMMDD'."""
    if raw is None:
        return ''
    s = str(raw).strip().replace('-', '')
    if len(s) == 8 and s.isdigit():
        return s
    return ''


# ── Hovedlogikk ───────────────────────────────────────────────────────────────

def main():
    print('=' * 55)
    print('Genererer Borregaard_SA_Master_v2.xlsx  (FASE 7.2)')
    print('=' * 55)

    # ── 1. data_7.xlsx — primærkilde ──────────────────────────────────────────
    print('Leser data_7.xlsx (primærkilde)...')
    data7_rows = read_sheet(DATA7_FILE)
    print(f'  {len(data7_rows)} rader')

    data7_by_varenr = {}
    for row in data7_rows:
        varenr = val(row, 'VareNr')
        if not varenr:
            continue
        key = str(varenr).strip()
        if key not in data7_by_varenr:
            data7_by_varenr[key] = row

    sa_nr_by_varenr    = {}
    lokasjon_by_varenr = {}
    varestatus_by_art  = {}
    erstatning_by_art  = {}

    for varenr, row in data7_by_varenr.items():
        sa_nr    = val(row, 'Kundens artnr')
        lokasjon = val(row, 'Kundens artbeskr.')
        vs       = val(row, 'VareStatus', 'Varestatus')
        erst     = val(row, 'ErsattsAvArtNr')
        alt      = val(row, 'Alternativ(er)')

        if sa_nr:
            sa_nr_by_varenr[varenr] = str(sa_nr).strip()

        # Aksepter ALLE ikke-tomme lokasjoner fra data_7 — inkl. BORSKUFF, LAGER1, tall osv.
        lok_str = str(lokasjon).strip() if lokasjon else ''
        if lok_str:
            lokasjon_by_varenr[varenr] = lok_str

        if vs:
            varestatus_by_art[varenr] = str(vs).strip()

        best = str(erst).strip() if erst and str(erst).strip() not in ('0', '') else ''
        if not best:
            best = str(alt).strip() if alt and str(alt).strip() not in ('0', '') else ''
        if best and best != varenr:
            erstatning_by_art[varenr] = best

    sa_dekning  = sum(1 for v in data7_by_varenr if v in sa_nr_by_varenr)
    lok_dekning = sum(1 for v in data7_by_varenr if v in lokasjon_by_varenr)
    print(f'  SA-nr:    {sa_dekning} av {len(data7_by_varenr)} artikler')
    print(f'  Lokasjon: {lok_dekning} av {len(data7_by_varenr)} artikler')
    print(f'  VareStatus: {len(varestatus_by_art)} artikler')
    print(f'  Erstatning: {len(erstatning_by_art)} artikler')

    # ── 2. SA-Nummer.xlsx — fallback SA-nr kun ────────────────────────────────
    print('Leser SA-Nummer.xlsx (fallback)...')
    sa_rows = read_sheet(SA_FILE)
    print(f'  {len(sa_rows)} rader')

    fallback_sa_count = 0
    for row in sa_rows:
        varenr = val(row, 'Artikelnr')
        sa_nr  = val(row, 'Kunds artikkelnummer', 'Kunds artikelnummer')
        if not varenr:
            continue
        key = str(varenr).strip()
        if sa_nr and key not in sa_nr_by_varenr:
            sa_nr_by_varenr[key] = str(sa_nr).strip()
            fallback_sa_count += 1

    print(f'  Fallback SA-nr lagt til:   {fallback_sa_count}')
    print(f'  Totalt SA-nr-oppslag: {len(sa_nr_by_varenr)}')

    # ── 3. Master_Artikkelstatus.xlsx ─────────────────────────────────────────
    print('Leser Master_Artikkelstatus.xlsx...')
    master_rows = read_sheet(MASTER_FILE)
    print(f'  {len(master_rows)} rader')
    master_by_art = {str(val(r, 'Artikelnr')).strip(): r for r in master_rows if val(r, 'Artikelnr')}

    # ── 4. Analyse_Lagerplan.xlsx ─────────────────────────────────────────────
    print('Leser Analyse_Lagerplan.xlsx...')
    lagerplan_rows = read_sheet(LAGERPLAN_FILE)
    print(f'  {len(lagerplan_rows)} rader')
    lagerplan_by_art = {str(val(r, 'Artikelnr')).strip(): r for r in lagerplan_rows if val(r, 'Artikelnr')}
    print(f'  Master-oppslag: {len(master_by_art)} | Lagerplan-oppslag: {len(lagerplan_by_art)}')

    # ── 5. Master.xlsx — fallback erstatning + InvDat ─────────────────────────
    print('Leser Master.xlsx (ErsattsAvArtNr fallback + InvDat)...')
    master_full_rows = read_sheet(MASTER_FULL_FILE)
    print(f'  {len(master_full_rows)} rader')

    invdat_by_art = {}
    master_fallback_count = 0

    for row in master_full_rows:
        art    = val(row, 'Artikelnr')
        erst   = val(row, 'Ersätts av artikel')
        invdat = row.get('InvDat')
        if not art:
            continue
        art_key = str(art).strip()
        erst_str = str(erst).strip() if erst else ''
        if erst_str and erst_str not in ('0', '') and erst_str != art_key and art_key not in erstatning_by_art:
            erstatning_by_art[art_key] = erst_str
            master_fallback_count += 1
        formatted = format_invdat(invdat)
        if formatted:
            invdat_by_art[art_key] = formatted

    print(f'  Master fallback erstatning: {master_fallback_count}')
    print(f'  Totalt erstatnings-oppslag: {len(erstatning_by_art)}')
    print(f'  InvDat-oppslag: {len(invdat_by_art)} artikler')

    # ── 6. leverandører.xlsx ──────────────────────────────────────────────────
    print('Leser leverandører.xlsx...')
    lev_rows = read_sheet(LEVERANDOR_FILE)
    print(f'  {len(lev_rows)} rader')

    lev_ledtid_map = {}
    for row in lev_rows:
        nr = val(row, 'Företagsnr')
        if not nr:
            continue
        key = str(nr).strip()
        ledtid    = val(row, 'LevLedTid',     default=None)
        transport = val(row, 'Transportdagar', default=None)
        if key not in lev_ledtid_map:
            try:
                lev_val = int(float(str(ledtid).replace(',', '.')))    if ledtid    not in (None, '') else 0
                tra_val = int(float(str(transport).replace(',', '.'))) if transport not in (None, '') else 0
            except (ValueError, TypeError):
                lev_val, tra_val = 0, 0
            lev_ledtid_map[key] = (lev_val, tra_val)

    print(f'  Ledetid-oppslag: {len(lev_ledtid_map)} leverandører')

    # ── 7. Ordrer_Jeeves.xlsx — salgshistorikk ────────────────────────────────
    ordre_by_art = {}
    if os.path.exists(ORDRER_FILE):
        print('Leser Ordrer_Jeeves.xlsx (salgshistorikk)...')
        orders_rows = read_sheet(ORDRER_FILE)
        print(f'  {len(orders_rows)} rader')
        for row in orders_rows:
            item_id = val(row, 'Item ID')
            if not item_id:
                continue
            key = str(item_id).strip()
            qty   = row.get('Delivered quantity') or 0
            verdi = row.get('Delivered value')   or 0
            dato  = val(row, 'Date')
            if key not in ordre_by_art:
                ordre_by_art[key] = {'antall': 0, 'verdi': 0, 'siste': '', 'linjer': 0}
            try:
                ordre_by_art[key]['antall'] += float(str(qty).replace(',', '.'))
                ordre_by_art[key]['verdi']  += float(str(verdi).replace(',', '.'))
                ordre_by_art[key]['linjer'] += 1
            except (ValueError, TypeError):
                pass
            dato_str = str(dato).strip()[:10] if dato else ''
            if dato_str > ordre_by_art[key]['siste']:
                ordre_by_art[key]['siste'] = dato_str
        print(f'  Ordre-oppslag: {len(ordre_by_art)} artikler med salgsdata')
    else:
        print(f'  ADVARSEL: Ordrer_Jeeves.xlsx ikke funnet — Ordre_TotAntall/Verdi/SisteDato/Antall forblir tomme')

    # ── Kolonner i output ──────────────────────────────────────────────────────
    output_columns = [
        'SA_Nummer', 'Tools_ArtNr', 'Beskrivelse',
        'Lagersaldo', 'DispLagSaldo', 'BP', 'Maxlager', 'ReservAnt', 'BestAntLev',
        'R12 Del Qty', 'Artikelstatus', 'Supplier Name', 'Lagerhylla',
        'VareStatus', 'ErsattsAvArtNr',
        'LAGERFØRT', 'VAREMERKE', 'PakkeStørrelse', 'Enhet / Ant Des',
        'Item category 1', 'Item category 2', 'Item category 3',
        'Ordre_TotAntall', 'Ordre_TotVerdi', 'Ordre_SisteDato', 'Ordre_Antall',
        'Dagens_Pris',
        'Kalkylpris_bas', 'EOK',
        'LevLedTid', 'Transportdagar',
        'InvDat',
    ]

    # ── Generer output-rader ───────────────────────────────────────────────────
    output_rows = []
    skipped = 0

    for varenr, row in data7_by_varenr.items():
        sa_nr = sa_nr_by_varenr.get(varenr, '')
        if not sa_nr:
            skipped += 1
            continue

        m = master_by_art.get(varenr, {})
        l = lagerplan_by_art.get(varenr, {})

        # Prioritert lokasjonskilde:
        # 1. Kundens artbeskr. fra data_7 (allerede i lokasjon_by_varenr)
        # 2. Hylla 1 fra Master_Artikkelstatus
        # 3. Lagerhylla fra Master_Artikkelstatus
        # 4. Tom streng
        lokasjon = lokasjon_by_varenr.get(varenr, '')
        if not lokasjon:
            lokasjon = str(m.get('Hylla 1', '') or '').strip()
        if not lokasjon or lokasjon == 'nan':
            lokasjon = str(val(m, 'Lagerhylla') or '').strip()
        if lokasjon == 'nan':
            lokasjon = ''
        varestatus = varestatus_by_art.get(varenr, '')
        erstatning = erstatning_by_art.get(varenr, '')
        kalkylpris = val(m, 'Kalkylpris bas_2', 'Kalkylpris bas')
        invdat     = invdat_by_art.get(varenr, '')

        supplier_nr  = val(l, 'Leverantör')
        supplier_key = str(supplier_nr).strip() if supplier_nr else ''
        lev_tuple    = lev_ledtid_map.get(supplier_key, (0, 0))

        beskrivelse = val(row, 'Artikelbeskrivning') or val(m, 'Artikelbeskrivning')

        o = ordre_by_art.get(varenr, {})

        output_rows.append({
            'SA_Nummer':       sa_nr,
            'Tools_ArtNr':     varenr,
            'Beskrivelse':     beskrivelse,
            'Lagersaldo':      val(m, 'TotLagSaldo', 'Lagersaldo'),
            'DispLagSaldo':    val(m, 'DispLagSaldo'),
            'BP':              val(l, 'BP'),
            'Maxlager':        val(l, 'Maxlager'),
            'ReservAnt':       val(m, 'ReservAnt'),
            'BestAntLev':      val(m, 'BestAntLev'),
            'R12 Del Qty':     '',
            'Artikelstatus':   val(m, 'Artikelstatus'),
            'Supplier Name':   val(l, 'Leverantör'),
            'Lagerhylla':      lokasjon,
            'VareStatus':      varestatus,
            'ErsattsAvArtNr':  erstatning,
            'LAGERFØRT':       val(row, 'LAGERFØRT'),
            'VAREMERKE':       val(row, 'VAREMERKE'),
            'PakkeStørrelse':  val(row, 'PakkeStørrelse'),
            'Enhet / Ant Des': val(row, 'Enhet / Ant Des'),
            'Item category 1': val(m, 'Varugrupp'),
            'Item category 2': val(row, 'NordicCategoryStruct5'),
            'Item category 3': '',
            'Ordre_TotAntall': int(o.get('antall', 0)) or '',
            'Ordre_TotVerdi':  round(o.get('verdi', 0)) or '',
            'Ordre_SisteDato': o.get('siste', ''),
            'Ordre_Antall':    o.get('linjer', 0) or '',
            'Dagens_Pris':     '',
            'Kalkylpris_bas':  kalkylpris,
            'EOK':             val(l, 'EOK'),
            'LevLedTid':       lev_tuple[0],
            'Transportdagar':  lev_tuple[1],
            'InvDat':          invdat,
        })

    # ── Skriv output ──────────────────────────────────────────────────────────
    print(f'\nSkriver {OUTPUT_FILE}...')
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = 'SA-Oversikt'
    ws_out.append(output_columns)
    for row_dict in output_rows:
        ws_out.append([row_dict.get(col, '') for col in output_columns])
    wb_out.save(OUTPUT_FILE)

    # ── Verifisering ──────────────────────────────────────────────────────────
    erstatning_count = sum(1 for r in output_rows if r.get('ErsattsAvArtNr'))
    ledetid_count    = sum(1 for r in output_rows if r.get('LevLedTid', 0) > 0 or r.get('Transportdagar', 0) > 0)
    invdat_count     = sum(1 for r in output_rows if r.get('InvDat'))
    telt2026_count   = sum(1 for r in output_rows if r.get('InvDat', '') >= '20260101' and r.get('InvDat'))
    varestatus_count = sum(1 for r in output_rows if r.get('VareStatus'))
    lagerfort_count  = sum(1 for r in output_rows if r.get('LAGERFØRT'))
    varemerke_count  = sum(1 for r in output_rows if r.get('VAREMERKE'))
    lok_count        = sum(1 for r in output_rows if r.get('Lagerhylla'))

    print(f'\nFerdig!')
    print(f'  Fil:      {OUTPUT_FILE}')
    print(f'  Ark:      "SA-Oversikt"')
    print(f'  Rader:    {len(output_rows)} artikler')
    print(f'  Hoppet over (ingen SA-nr): {skipped}')
    print(f'  Kolonner: {len(output_columns)}')
    print(f'\n  ── Verifisering ──')
    print(f'  Lagerhylla ikke-tomme:     {lok_count} av {len(output_rows)}')
    print(f'  ErsattsAvArtNr ikke-tomme: {erstatning_count} av {len(output_rows)}')
    print(f'  VareStatus ikke-tomme:     {varestatus_count} av {len(output_rows)}')
    print(f'  Artikler med ledetid:      {ledetid_count} av {len(output_rows)}')
    print(f'  InvDat (sist telt):        {invdat_count} av {len(output_rows)}')
    print(f'  Telt i 2026:               {telt2026_count} av {len(output_rows)}')
    print(f'  LAGERFØRT ikke-tomme:      {lagerfort_count} av {len(output_rows)}')
    print(f'  VAREMERKE ikke-tomme:      {varemerke_count} av {len(output_rows)}')


if __name__ == '__main__':
    main()